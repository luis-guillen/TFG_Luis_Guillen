import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font
import re
import argparse
from pathlib import Path
import logging

# Configuración de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('proceso_excel.log'),
        logging.StreamHandler()
    ]
)

def add_year_column(df, year):
    """
    Añade una columna 'year' después de la columna 'Código NIF' 
    con el valor del año especificado.
    
    Args:
        df (pd.DataFrame): DataFrame con los datos
        year (int): Año a añadir en la columna
        
    Returns:
        pd.DataFrame: DataFrame con la columna 'year' añadida
    """
    try:
        # Encontrar el índice de la columna 'Código NIF'
        nif_idx = df.columns.get_loc("Código NIF")
        
        # Insertar la columna 'year' después de 'Código NIF'
        df.insert(nif_idx + 1, "year", year)
        
        return df
        
    except Exception as e:
        logging.error(f"Error al añadir columna 'year': {str(e)}")
        return None


def consolidate_files(input_files, output_file):
    """
    Consolida múltiples archivos de Excel en uno solo,
    concatenando los datos de la hoja 'Resultados'.
    """
    print(f"Consolidando archivos en {output_file}...")
    
    if not input_files:
        print("Error: No se proporcionaron archivos de entrada.")
        return False
    
    try:
        # Crear un nuevo libro y copiar el primer archivo
        first_file = input_files[0]
        wb_source = load_workbook(first_file)
        ws_source = wb_source["Resultados"]
        
        wb_target = Workbook()
        # Eliminar la hoja por defecto
        default_sheet = wb_target.active
        wb_target.remove(default_sheet)
        
        # Copiar las hojas relevantes
        wb_target.create_sheet("Estrategia de búsqueda")
        ws_target = wb_target.create_sheet("Resultados")
        
        # Copiar datos de la primera hoja
        for row in ws_source.rows:
            values = [cell.value for cell in row]
            ws_target.append(values)
        
        # Altura de la última fila con datos
        last_row = ws_target.max_row
        
        # Procesar los archivos restantes
        for file in input_files[1:]:
            print(f"Añadiendo datos de {file}...")
            wb = load_workbook(file)
            ws = wb["Resultados"]
            
            # Añadir dos filas en blanco
            last_row += 2
            
            # Copiar datos, incluyendo encabezados
            for row_idx, row in enumerate(ws.rows, 1):
                values = [cell.value for cell in row]
                ws_target.append(values)
                last_row += 1
        
        # Guardar el archivo consolidado
        wb_target.save(output_file)
        print(f"Archivos consolidados correctamente en {output_file}")
        return True
    
    except Exception as e:
        print(f"Error al consolidar archivos: {str(e)}")
        return False


def format_consolidated_file(input_file, output_file):
    """
    Formatea el archivo consolidado: elimina filas en blanco,
    encabezados duplicados y reemplaza 'n.d.' por espacios en blanco.
    """
    print(f"Formateando archivo consolidado en {output_file}...")
    
    try:
        # Cargar el archivo
        wb = load_workbook(input_file)
        ws = wb["Resultados"]
        
        # Crear un nuevo libro
        wb_new = Workbook()
        default_sheet = wb_new.active
        wb_new.remove(default_sheet)
        ws_new = wb_new.create_sheet("Resultados")
        
        # Copiar el encabezado (primera fila)
        header_row = next(ws.rows)
        header_values = [cell.value for cell in header_row]
        ws_new.append(header_values)
        
        # Procesar las filas restantes
        skip_row = False
        for row_idx, row in enumerate(ws.rows, 1):
            # Saltar la primera fila (ya copiada)
            if row_idx == 1:
                continue
                
            # Detectar si es una fila de encabezado (duplicada)
            is_header = False
            cell_values = [cell.value for cell in row]
            
            # Si la primera celda coincide con el encabezado, asumimos que es un encabezado duplicado
            if cell_values and cell_values[0] == header_values[0]:
                is_header = True
            
            # Verificar si es una fila en blanco
            is_blank = all(cell.value is None or cell.value == "" for cell in row)
            
            # Si no es un encabezado duplicado ni una fila en blanco, copiar la fila
            if not is_header and not is_blank:
                # Reemplazar 'n.d.' por espacios en blanco
                processed_values = []
                for cell in row:
                    value = cell.value
                    if value == "n.d.":
                        processed_values.append(None)
                    else:
                        processed_values.append(value)
                
                ws_new.append(processed_values)
        
        # Aplicar formato: color de letra negro y celdas sin relleno
        black_font = Font(color="00000000")
        for row in ws_new.rows:
            for cell in row:
                cell.font = black_font
                cell.fill = PatternFill(fill_type=None)
        
        # Guardar el archivo formateado
        wb_new.save(output_file)
        print(f"Archivo formateado correctamente en {output_file}")
        return True
    
    except Exception as e:
        print(f"Error al formatear archivo: {str(e)}")
        return False


def group_by_company(input_file, output_file, headers_template):
    """
    Agrupa los datos por empresa y añade los encabezados estandarizados.
    """
    print(f"Agrupando datos por empresa en {output_file}...")
    
    try:
        # Cargar el archivo
        df = pd.read_excel(input_file, sheet_name="Resultados")
        
        # Ordenar por Nombre y luego por año
        df_sorted = df.sort_values(by=["Nombre", "year"])
        
        # Cargar plantilla de encabezados
        headers_df = pd.read_excel(headers_template)
        headers_rows = headers_df.iloc[0:3].copy()
        
        # Crear un nuevo libro con los datos ordenados
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Guardar los encabezados estandarizados y los datos ordenados
            pd.concat([headers_rows, df_sorted]).to_excel(writer, sheet_name="Resultados", index=False)
        
        print(f"Datos agrupados correctamente en {output_file}")
        return True
    
    except Exception as e:
        print(f"Error al agrupar datos: {str(e)}")
        return False


def transfer_to_bam_template(consolidated_file, template_file, company_cif, company_name):
    """
    Copia los datos de una empresa específica a la plantilla BAM.
    """
    output_file = f"BAM_{company_cif} ({company_name}).xlsx"
    print(f"Transfiriendo datos a la plantilla BAM para {company_name} (CIF: {company_cif})...")
    
    try:
        # Cargar el archivo consolidado
        df = pd.read_excel(consolidated_file, sheet_name="Resultados")
        
        # Filtrar por el CIF de la empresa
        company_data = df[df["Código CIF"] == company_cif]
        
        if company_data.empty:
            print(f"Error: No se encontraron datos para la empresa con CIF {company_cif}")
            return False
        
        # Verificar que los datos estén ordenados por año
        company_data = company_data.sort_values(by="year")
        
        # Seleccionar solo las columnas de datos contables (desde la columna H hasta el final)
        # Suponemos que las columnas de datos contables empiezan después de algunas columnas iniciales
        # Según el documento, parece que comienzan en la columna H (índice 7)
        accounting_data = company_data.iloc[:, 7:]
        
        # Copiar la plantilla BAM
        import shutil
        shutil.copy2(template_file, output_file)
        
        # Cargar la plantilla copiada
        wb = load_workbook(output_file)
        ws = wb["Carga_datos"]
        
        # Transponer los datos
        transposed_data = accounting_data.T
        
        # Escribir los datos transpuestos en la hoja "Carga_datos" a partir de la celda D2
        for i, (_, row) in enumerate(transposed_data.iterrows(), 1):
            for j, value in enumerate(row, 1):
                ws.cell(row=i+1, column=j+3, value=value)
        
        # Guardar el archivo
        wb.save(output_file)
        print(f"Datos transferidos correctamente a {output_file}")
        return True
    
    except Exception as e:
        print(f"Error al transferir datos a la plantilla BAM: {str(e)}")
        return False


def process_all_companies(consolidated_file, template_file):
    """
    Procesa todas las empresas del archivo consolidado y crea un archivo BAM para cada una.
    """
    print(f"Procesando todas las empresas del archivo {consolidated_file}...")
    
    try:
        # Cargar el archivo consolidado
        df = pd.read_excel(consolidated_file, sheet_name="Resultados")
        
        # Obtener lista única de empresas (CIF y Nombre)
        companies = df[["Código CIF", "Nombre"]].drop_duplicates()
        
        success_count = 0
        for _, company in companies.iterrows():
            cif = company["Código CIF"]
            name = company["Nombre"]
            
            # Limpiar el nombre para usarlo en el nombre de archivo
            clean_name = re.sub(r'[\\/*?:"<>|]', "", name)
            
            if transfer_to_bam_template(consolidated_file, template_file, cif, clean_name):
                success_count += 1
        
        print(f"Procesadas {success_count} de {len(companies)} empresas con éxito.")
        return success_count == len(companies)
    
    except Exception as e:
        print(f"Error al procesar todas las empresas: {str(e)}")
        return False


def process_all_files(input_dir, start_year=2008, end_year=2023):
    """
    Procesa todos los archivos Excel en el directorio de entrada,
    añadiendo la columna 'year' a cada uno.
    
    Args:
        input_dir (str): Directorio que contiene los archivos Excel
        start_year (int): Año inicial
        end_year (int): Año final
        
    Returns:
        dict: Diccionario con los DataFrames procesados, clave: año
    """
    try:
        input_dir = Path(input_dir)
        if not input_dir.exists():
            logging.error(f"El directorio {input_dir} no existe")
            return None
            
        processed_dfs = {}
        for year in range(start_year, end_year + 1):
            file_name = f"SABI_Export_{year}.xlsx"
            file_path = input_dir / file_name
            
            if not file_path.exists():
                logging.warning(f"Archivo no encontrado: {file_path}")
                continue
                
            logging.info(f"Procesando archivo: {file_path}")
            
            try:
                # Leer el archivo Excel
                df = pd.read_excel(file_path, sheet_name="Resultados")
                
                # Añadir columna 'year'
                df = add_year_column(df, year)
                
                if df is not None:
                    processed_dfs[year] = df
                    logging.info(f"Archivo {file_name} procesado correctamente")
                    
            except Exception as e:
                logging.error(f"Error al procesar {file_name}: {str(e)}")
                continue
                
        return processed_dfs
        
    except Exception as e:
        logging.error(f"Error en process_all_files: {str(e)}")
        return None


def main():
    # Directorio base del proyecto
    BASE_DIR = Path(__file__).parent.parent
    
    # Directorios y archivos
    input_dir = BASE_DIR / "Data" / "modificados"
    
    # Procesar todos los archivos
    processed_dfs = process_all_files(input_dir)
    
    if processed_dfs:
        logging.info(f"Procesados {len(processed_dfs)} archivos exitosamente")
        
        # Guardar los DataFrames procesados
        for year, df in processed_dfs.items():
            output_file = input_dir / f"SABI_Export_{year}_processed.xlsx"
            df.to_excel(output_file, sheet_name="Resultados", index=False)
            logging.info(f"Archivo guardado: {output_file}")
        return True
    else:
        logging.error("Hubo errores durante el proceso")
        return False


if __name__ == "__main__":
    if main():
        logging.info("Proceso completado exitosamente")
    else:
        logging.error("Hubo errores durante el proceso")