import pandas as pd
import os
import re
from openpyxl import load_workbook

def add_year_column(df, year):
    """Añade columna de año después de la columna 'Código NIF'"""
    nif_col_idx = df.columns.get_loc('Código NIF')
    df.insert(nif_col_idx + 1, 'year', year)
    return df

def clean_nd_values(df):
    """Reemplaza valores 'n.d.' con cadenas vacías"""
    return df.replace('n.d.', '')

def main():
    # Configuración
    start_year = 2008
    end_year = 2023
    base_dir = 'Data/modificados'  # Directorio donde están los datos
    
    print(f"Iniciando proceso de consolidación de datos BAM para los años {start_year}-{end_year}")
    
    # Paso 0: Procesar cada archivo anual
    all_data = []
    
    for year in range(start_year, end_year + 1):
        input_file = os.path.join(base_dir, f"SABI_Export_{year}.xlsx")  # Nombre de archivo corregido
        if not os.path.exists(input_file):
            print(f"Advertencia: Archivo {input_file} no encontrado. Se omite el año {year}.")
            continue
            
        print(f"Procesando archivo del año {year}: {input_file}")
        
        # Leer la hoja "Resultados"
        df = pd.read_excel(input_file, sheet_name="Resultados")
        
        # Mostrar información sobre los datos
        print(f"  Columnas encontradas: {df.columns.tolist()}")
        print(f"  Número de filas: {len(df)}")
        
        # Añadir columna de año
        df = add_year_column(df, year)
        
        # Añadir a la lista de todos los dataframes
        all_data.append(df)
    
    if not all_data:
        print("No se encontraron archivos de datos. Por favor, verifique las rutas de los archivos.")
        return
    
    # Paso 1: Consolidación
    print("\nPaso 1: Consolidando datos...")
    consolidated_df = pd.concat(all_data, ignore_index=True)
    consolidated_file = os.path.join(base_dir, f"Sabi_Export_{start_year}_{end_year}_hoteles_paso_01.xlsx")
    consolidated_df.to_excel(consolidated_file, sheet_name="Resultados", index=False)
    print(f"Datos consolidados guardados en {consolidated_file}")
    print(f"  Número total de filas: {len(consolidated_df)}")
    print(f"  Columnas en el archivo consolidado: {consolidated_df.columns.tolist()}")
    
    # Paso 2: Formateo
    print("\nPaso 2: Formateando archivo consolidado...")
    # Limpiar valores n.d.
    formatted_df = clean_nd_values(consolidated_df)
    formatted_file = os.path.join(base_dir, f"Sabi_Export_{start_year}_{end_year}_hoteles_paso_02.xlsx")
    formatted_df.to_excel(formatted_file, sheet_name="Resultados", index=False)
    print(f"Datos formateados guardados en {formatted_file}")
    
    # Paso 3: Agrupación por empresa
    print("\nPaso 3: Agrupando datos por empresa...")
    # Ordenar por nombre de empresa y luego por año
    grouped_df = formatted_df.sort_values(by=["Nombre", "year"])
    
    # Añadir encabezados estandarizados si el archivo existe
    headers_file = os.path.join(base_dir, "0_Plantilla_encabezados_estandarizados.xlsx")
    if os.path.exists(headers_file):
        print("Añadiendo encabezados estandarizados...")
        headers_df = pd.read_excel(headers_file)
        headers = headers_df.iloc[:3].values.tolist()
        
        # Crear un nuevo archivo Excel con los encabezados
        grouped_file = os.path.join(base_dir, f"Sabi_Export_{start_year}_{end_year}_hoteles_paso_03.xlsx")
        grouped_df.to_excel(grouped_file, sheet_name="Resultados", index=False, startrow=3)
        
        # Usar openpyxl para añadir los encabezados
        workbook = load_workbook(grouped_file)
        sheet = workbook["Resultados"]
        
        # Añadir encabezados
        for i, header_row in enumerate(headers):
            for j, value in enumerate(header_row):
                if pd.notna(value):  # Solo añadir valores no-NaN
                    sheet.cell(row=i+1, column=j+1, value=value)
        
        workbook.save(grouped_file)
    else:
        print("Advertencia: Archivo de encabezados estandarizados no encontrado.")
        grouped_file = os.path.join(base_dir, f"Sabi_Export_{start_year}_{end_year}_hoteles_paso_03.xlsx")
        grouped_df.to_excel(grouped_file, sheet_name="Resultados", index=False)
    
    print(f"Datos agrupados guardados en {grouped_file}")
    
    # Paso 4: Transferencia a plantilla BAM
    print("\nPaso 4: Preparando transferencia a plantilla BAM...")
    
    # Obtener empresas únicas
    unique_companies = grouped_df[["Código NIF", "Nombre"]].drop_duplicates()
    print(f"Se encontraron {len(unique_companies)} empresas únicas.")
    if len(unique_companies) > 0:
        print("  Ejemplo de empresas encontradas:")
        print(unique_companies.head())
    
    template_file = os.path.join(base_dir, "1_Medium_BAM_plantilla_10_años_copia.xlsx")
    if not os.path.exists(template_file):
        print(f"Advertencia: Archivo de plantilla BAM {template_file} no encontrado. No se puede continuar con el Paso 4.")
        return
    
    # Procesar cada empresa
    for idx, (nif, name) in enumerate(zip(unique_companies["Código NIF"], unique_companies["Nombre"])):
        print(f"Procesando empresa {idx+1}/{len(unique_companies)}: {name} (NIF: {nif})")
        
        # Limpiar el nombre de la empresa para el nombre del archivo
        clean_name = re.sub(r'[\\/*?:"<>|]', "", name)  # Eliminar caracteres inválidos para nombres de archivo
        bam_file = os.path.join(base_dir, f"BAM_{nif} ({clean_name}).xlsx")
        
        # Copiar la plantilla
        template_wb = load_workbook(template_file)
        template_wb.save(bam_file)
        
        # Obtener datos de la empresa
        company_data = grouped_df[grouped_df["Código NIF"] == nif].sort_values(by="year")
        
        if len(company_data) == 0:
            print(f"  Advertencia: No se encontraron datos para la empresa {name}. Se omite.")
            continue
            
        # Preparar los datos para la transposición
        # Asumir que las columnas de datos comienzan desde la columna H (índice 7) y continúan hasta el final
        data_cols = company_data.columns[7:]
        company_data_subset = company_data[data_cols]
        
        # Transponer los datos
        transposed_data = company_data_subset.T
        
        # Guardar en Excel
        with pd.ExcelWriter(bam_file, engine='openpyxl', mode='a') as writer:
            transposed_data.to_excel(writer, sheet_name="Carga_datos", startrow=1, startcol=3, header=False, index=True)
        
        print(f"  Archivo BAM creado: {bam_file}")
    
    print("\n¡Proceso completado con éxito!")

if __name__ == "__main__":
    main()